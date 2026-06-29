#!/usr/bin/env python3
"""Export the SSBM data sheet workbook into searchable CSV artifacts."""

from __future__ import annotations

import csv
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


SCRIPT = Path(__file__).resolve()
for parent in SCRIPT.parents:
    if (parent / "configure.py").exists() and (parent / "config" / "GALE01").exists():
        ROOT = parent
        break
else:
    raise RuntimeError("Could not find repo root from exporter path")

for parent in SCRIPT.parents:
    if parent.name == "decomp_resources":
        RESOURCE_DIR = parent
        break
else:
    raise RuntimeError("Could not find decomp_resources root from exporter path")

SLICE_DIR = RESOURCE_DIR / "data_sheets" / "ssbm_data_sheet_1_02"
SOURCE = SLICE_DIR / "source" / "ssbm_data_sheet_1_02.xlsx"
OUT_DIR = SLICE_DIR / "csv"

SHEET_DESCRIPTIONS = {
    "Global Addresses": (
        "Known global/static memory addresses with effects and notes. Useful for "
        "RAM watches, patch points, and locating data-backed behavior."
    ),
    "Function Addresses": (
        "Known function entry points with descriptions and register notes. Useful "
        "for debugger lookups and call-site interpretation."
    ),
    "Action State Reference": (
        "Move table and action-state structure offsets. Useful when mapping "
        "fighter state constants and animation/action metadata."
    ),
    "ID Lists": (
        "Mixed ID tables for characters, stages, CPU AI, items, elements, status "
        "flags, SFX, name tags, and music."
    ),
    "Stage Data Offsets": (
        "Static stage info structure offsets. Useful for camera bounds, spawn "
        "tables, collision metadata, and stage-specific constants."
    ),
    "Entity Data Offsets": (
        "GObj/entity structure layout notes. Useful for object ownership, links, "
        "callbacks, and entity lifecycle fields."
    ),
    "Hitbox Offsets": (
        "Hitbox structure offsets and field semantics. Useful for subaction event "
        "decoding and fighter/item collision behavior."
    ),
    "GmGover Offsets": (
        "Game-over object/view offsets with default hex/value/effect data."
    ),
    "Hurtbox Offsets": (
        "Hurtbox structure offset notes. Useful for collision body layout and "
        "fighter data interpretation."
    ),
    "PlCo Offsets": (
        "Player/common constant offsets and defaults. Useful for controller, "
        "movement, and shared gameplay tuning values."
    ),
    "Bones": (
        "Bone lookup table and body-part mapping notes. Useful for model joint "
        "lookups, hitbox attachment, and animation data."
    ),
    "StartMelee Struct": (
        "StartMelee match setup structure layout. Useful for VS setup, special "
        "mode setup, and match initialization data."
    ),
    "Char Data Offsets": (
        "Static and dynamic player/fighter data offsets. Useful for field naming, "
        "struct reconstruction, and debugger memory inspection."
    ),
    "Character Attributes": (
        "Per-character attribute table. Columns are attribute offsets; rows are "
        "characters. Useful for constants, float data, and character tuning."
    ),
    "Subaction Events": (
        "Subaction event binary layouts, including hitbox event bitfields. Useful "
        "for action script decoding and assembly-to-data interpretation."
    ),
    "MnSlChr Offsets": (
        "Character select screen offsets and default values."
    ),
    "Debug Menu Map": (
        "Debug menu memory map plus character and stage ID mappings."
    ),
    "Free Memory": (
        "Free or unused memory ranges with notes. Useful for patch experiments and "
        "understanding gaps in the address space."
    ),
}


def slugify(name: str, used: set[str]) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    slug = slug or "sheet"
    base = slug
    counter = 2
    while slug in used:
        slug = f"{base}_{counter}"
        counter += 1
    used.add(slug)
    return slug


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.15g}"
    text = str(value)
    text = text.replace("\r\n", "\\n").replace("\r", "\\n").replace("\n", "\\n")
    return text.strip()


def cell_text(value_cell: Any, formula_cell: Any) -> tuple[str, str]:
    formula_value = formula_cell.value
    formula = ""
    if isinstance(formula_value, str) and formula_value.startswith("="):
        formula = formula_value
    value = value_cell.value
    if value is None and formula:
        value = formula
    return clean_text(value), clean_text(formula)


def used_bounds(ws: Any) -> tuple[int, int, int, int] | None:
    min_row = min_col = max_row = max_col = None
    for row in ws.iter_rows():
        for cell in row:
            has_content = cell.value is not None or cell.hyperlink or cell.comment
            if not has_content:
                continue
            row_num = cell.row
            col_num = cell.column
            min_row = row_num if min_row is None else min(min_row, row_num)
            min_col = col_num if min_col is None else min(min_col, col_num)
            max_row = row_num if max_row is None else max(max_row, row_num)
            max_col = col_num if max_col is None else max(max_col, col_num)
    if min_row is None:
        return None
    return min_row, min_col, max_row, max_col


def nearest_above(labels_by_col: dict[int, str], col_num: int) -> str:
    return labels_by_col.get(col_num, "")


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    for old_csv in OUT_DIR.glob("*.csv"):
        old_csv.unlink()

    wb_values = openpyxl.load_workbook(SOURCE, data_only=True, read_only=False)
    wb_formulas = openpyxl.load_workbook(SOURCE, data_only=False, read_only=False)

    used_slugs: set[str] = set()
    sheet_index_rows: list[dict[str, str]] = []
    all_cell_rows: list[dict[str, str]] = []

    for ws_values, ws_formulas in zip(wb_values.worksheets, wb_formulas.worksheets):
        sheet_name = ws_values.title
        slug = slugify(sheet_name, used_slugs)
        sheet_csv_path = OUT_DIR / f"{slug}.csv"
        rel_sheet_csv = sheet_csv_path.relative_to(ROOT).as_posix()
        bounds = used_bounds(ws_formulas)

        if bounds is None:
            write_csv(sheet_csv_path, [], ["source_row", "row_text"])
            sheet_index_rows.append(
                {
                    "source_file": SOURCE.relative_to(ROOT).as_posix(),
                    "sheet_name": sheet_name,
                    "csv_path": rel_sheet_csv,
                    "used_range": "",
                    "rows": "0",
                    "columns": "0",
                    "non_empty_cells": "0",
                    "formula_cells": "0",
                    "hyperlink_cells": "0",
                    "comment_cells": "0",
                    "description": SHEET_DESCRIPTIONS.get(sheet_name, ""),
                }
            )
            continue

        min_row, min_col, max_row, max_col = bounds
        col_letters = [get_column_letter(col) for col in range(min_col, max_col + 1)]
        wide_fields = ["source_row", "row_text", *col_letters]
        wide_rows: list[dict[str, str]] = []
        labels_by_col: dict[int, str] = {}
        non_empty_cells = 0
        formula_cells = 0
        hyperlink_cells = 0
        comment_cells = 0

        for row_num in range(min_row, max_row + 1):
            row_values: list[str] = []
            wide_row = {"source_row": str(row_num)}
            row_has_content = False

            for col_num in range(min_col, max_col + 1):
                value_cell = ws_values.cell(row_num, col_num)
                formula_cell = ws_formulas.cell(row_num, col_num)
                value, formula = cell_text(value_cell, formula_cell)
                hyperlink = ""
                if formula_cell.hyperlink:
                    hyperlink = formula_cell.hyperlink.target or ""
                comment = clean_text(formula_cell.comment.text if formula_cell.comment else None)
                visible_value = value or hyperlink or comment
                if visible_value:
                    row_values.append(visible_value)
                    row_has_content = True
                wide_row[get_column_letter(col_num)] = value

            if not row_has_content:
                continue

            row_text = " | ".join(row_values)
            wide_row["row_text"] = row_text
            wide_rows.append(wide_row)

            row_label = row_values[0] if row_values else ""
            for col_num in range(min_col, max_col + 1):
                value_cell = ws_values.cell(row_num, col_num)
                formula_cell = ws_formulas.cell(row_num, col_num)
                value, formula = cell_text(value_cell, formula_cell)
                hyperlink = ""
                if formula_cell.hyperlink:
                    hyperlink = formula_cell.hyperlink.target or ""
                comment = clean_text(formula_cell.comment.text if formula_cell.comment else None)
                if not (value or formula or hyperlink or comment):
                    continue

                above_context = nearest_above(labels_by_col, col_num)
                if value:
                    non_empty_cells += 1
                if formula:
                    formula_cells += 1
                if hyperlink:
                    hyperlink_cells += 1
                if comment:
                    comment_cells += 1

                cell_addr = f"{get_column_letter(col_num)}{row_num}"
                all_cell_rows.append(
                    {
                        "source_file": SOURCE.relative_to(ROOT).as_posix(),
                        "sheet_name": sheet_name,
                        "sheet_csv": rel_sheet_csv,
                        "row": str(row_num),
                        "column": str(col_num),
                        "cell": cell_addr,
                        "value": value,
                        "formula": formula,
                        "hyperlink": hyperlink,
                        "comment": comment,
                        "row_label": row_label,
                        "nearest_above": above_context,
                        "row_text": row_text,
                    }
                )
                if value:
                    labels_by_col[col_num] = value

        write_csv(sheet_csv_path, wide_rows, wide_fields)
        sheet_index_rows.append(
            {
                "source_file": SOURCE.relative_to(ROOT).as_posix(),
                "sheet_name": sheet_name,
                "csv_path": rel_sheet_csv,
                "used_range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
                "rows": str(len(wide_rows)),
                "columns": str(max_col - min_col + 1),
                "non_empty_cells": str(non_empty_cells),
                "formula_cells": str(formula_cells),
                "hyperlink_cells": str(hyperlink_cells),
                "comment_cells": str(comment_cells),
                "description": SHEET_DESCRIPTIONS.get(sheet_name, ""),
            }
        )

    write_csv(
        OUT_DIR / "sheet_index.csv",
        sheet_index_rows,
        [
            "source_file",
            "sheet_name",
            "csv_path",
            "used_range",
            "rows",
            "columns",
            "non_empty_cells",
            "formula_cells",
            "hyperlink_cells",
            "comment_cells",
            "description",
        ],
    )
    write_csv(
        OUT_DIR / "cells.csv",
        all_cell_rows,
        [
            "source_file",
            "sheet_name",
            "sheet_csv",
            "row",
            "column",
            "cell",
            "value",
            "formula",
            "hyperlink",
            "comment",
            "row_label",
            "nearest_above",
            "row_text",
        ],
    )

    print(f"Wrote {len(sheet_index_rows)} sheet records")
    print(f"Wrote {len(all_cell_rows)} searchable cell records")
    print(f"Output: {OUT_DIR.relative_to(ROOT).as_posix()}")


if __name__ == "__main__":
    main()
